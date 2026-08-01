#!/usr/bin/env python3
"""
导出指定日期范围的 CFFEX 股指期货历史数据到 Excel。

用法:
    python export_history.py                           # 默认 2026-06-01 ~ 今天
    python export_history.py --start 2026-01-01 --end 2026-06-30
    python export_history.py --output ./my_data.xlsx
"""

import argparse
import sys
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INSTRUMENTS = ["IH", "IF", "IC", "IM"]
INSTRUMENT_NAMES = {"IH": "上证50", "IF": "沪深300", "IC": "中证500", "IM": "中证1000"}

# 合约探测范围：当前月份前后各 12 个月
MONTH_RANGE = 12

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=13)
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
RED_FILL = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")  # 贴水行
COLORS = {
    "IH": "4472C4",
    "IF": "ED7D31",
    "IC": "70AD47",
    "IM": "FFC000",
}


def _safe_float(val):
    """安全转 float，失败返回 None。"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val):
    """安全转 int，失败返回 None。"""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _has_akshare():
    try:
        import akshare  # noqa: F401
        return True
    except ImportError:
        return False


def gen_monthly_contracts(year, month, count):
    """生成连续月份合约代码（如 IH2608）。"""
    codes = []
    for i in range(count):
        m = month + i
        y = year + (m - 1) // 12
        m = (m - 1) % 12 + 1
        for sym in INSTRUMENTS:
            codes.append(f"{sym}{y % 100:02d}{m:02d}")
    return codes


def discover_active_contracts(min_last_date=None):
    """查询并返回活跃的合约代码列表。"""
    import akshare as ak
    from datetime import timedelta

    today = date.today()
    if min_last_date is None:
        min_last_date = (today - timedelta(days=90)).strftime("%Y-%m-%d")

    # 确定需要覆盖的合约月份范围：从 min_last_date 所在月到 today + MONTH_RANGE
    min_y, min_m = int(min_last_date[:4]), int(min_last_date[:7].split("-")[1])
    months_back = (today.year - min_y) * 12 + (today.month - min_m)
    total_months = months_back + MONTH_RANGE + 2  # +2 留裕量
    candidates = gen_monthly_contracts(min_y, min_m, total_months)

    active = set()
    for code in candidates:
        try:
            df = ak.futures_zh_daily_sina(symbol=code)
            if df is None or df.empty:
                continue
            last_date = str(df.iloc[-1]["date"])
            if last_date >= min_last_date:
                active.add(code)
        except Exception:
            pass

    # 排序：IH → IF → IC → IM，同品种按月份
    instr_order = {"IH": 0, "IF": 1, "IC": 2, "IM": 3}
    return sorted(active, key=lambda c: (instr_order.get(c[:2], 99), c[2:]))


def fetch_all_history(contracts, start_date, end_date):
    """拉取所有合约在日期范围内的日线数据（含当日实时补丁）。"""
    import akshare as ak
    import pandas as pd

    RT_SYMBOLS = {"IH": "上证50指数期货", "IF": "沪深300指数期货", "IC": "中证500指数期货", "IM": "中证1000股指期货"}

    all_data = {}  # {contract_code: DataFrame}
    for code in contracts:
        try:
            df = ak.futures_zh_daily_sina(symbol=code)
            if df is None or df.empty:
                print(f"  [SKIP] {code}: 无数据")
                continue
            df["date"] = df["date"].astype(str)
            mask = (df["date"] >= start_date) & (df["date"] <= end_date)
            df = df[mask].copy()
            if not df.empty:
                all_data[code] = df
                print(f"  [OK] {code}: {len(df)} 条记录 ({df['date'].iloc[0]} ~ {df['date'].iloc[-1]})")
            else:
                print(f"  [SKIP] {code}: 日期范围内无数据")
        except Exception as e:
            print(f"  [ERR] {code}: {e}")

    # 补充当日数据：futures_zh_daily_sina 延迟 1-2 小时，用 realtime 补
    try:
        for inst, cn_name in RT_SYMBOLS.items():
            try:
                rt = ak.futures_zh_realtime(symbol=cn_name)
                if rt is None or rt.empty:
                    continue
                for _, row in rt.iterrows():
                    symbol = str(row.get("symbol", ""))
                    if symbol.endswith("0") or len(symbol) <= 4:
                        continue
                    if symbol not in all_data:
                        continue
                    trade_date = str(row.get("tradedate", ""))
                    if not trade_date:
                        continue
                    existing_dates = set(all_data[symbol]["date"].values)
                    if trade_date in existing_dates:
                        continue

                    close_v = _safe_float(row.get("close"))
                    volume_v = _safe_int(row.get("volume"))
                    if close_v is None or volume_v is None or volume_v == 0:
                        continue

                    new_row = pd.DataFrame([{
                        "date": trade_date,
                        "open": _safe_float(row.get("open")),
                        "high": _safe_float(row.get("high")),
                        "low": _safe_float(row.get("low")),
                        "close": close_v,
                        "volume": volume_v,
                        "hold": _safe_int(row.get("position")),
                        "settle": _safe_float(row.get("settlement")),
                    }])
                    all_data[symbol] = pd.concat([all_data[symbol], new_row], ignore_index=True)
                    print(f"  [+] {symbol}: 补充当日 {trade_date} (realtime)")
            except Exception:
                pass
    except Exception:
        pass

    return all_data


def build_summary_sheet(wb, all_data, start_date, end_date):
    """构建汇总 sheet：每个品种一行，列=日期，值=当月合约收盘价。"""
    ws = wb.active
    ws.title = "当月合约收盘价"

    # 收集所有交易日
    all_dates = set()
    for df in all_data.values():
        all_dates.update(df["date"].tolist())
    all_dates = sorted(all_dates)

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_dates) + 1)
    title_cell = ws.cell(row=1, column=1, value=f"CFFEX 股指期货当月合约收盘价  |  {start_date} ~ {end_date}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # 表头
    headers = ["品种"] + all_dates
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 每品种找当月合约
    for row_idx, inst in enumerate(INSTRUMENTS, 3):
        ws.cell(row=row_idx, column=1, value=f"{inst} ({INSTRUMENT_NAMES[inst]})").font = DATA_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER

        # 找该品种最早到期月份
        inst_codes = sorted(
            [c for c in all_data if c.startswith(inst)],
            key=lambda c: c[2:],
        )
        if not inst_codes:
            continue
        front_code = inst_codes[0]
        df = all_data[front_code].set_index("date")

        for col, d in enumerate(all_dates, 2):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")
            if d in df.index:
                cell.value = round(float(df.loc[d, "close"]), 1)

    # 列宽
    ws.column_dimensions["A"].width = 18
    for col in range(2, len(all_dates) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 10

    return ws


def build_detail_sheet(wb, all_data, inst):
    """构建单品种明细 sheet：行为日期，列为各合约 OHLCV。"""
    ws = wb.create_sheet(title=f"{inst} 全合约")

    inst_codes = sorted([c for c in all_data if c.startswith(inst)], key=lambda c: c[2:])
    if not inst_codes:
        ws.cell(row=1, column=1, value="无数据")
        return ws

    # 收集所有日期
    all_dates = set()
    for code in inst_codes:
        all_dates.update(all_data[code]["date"].tolist())
    all_dates = sorted(all_dates)

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(inst_codes) * 5 + 1)
    title_cell = ws.cell(row=1, column=1, value=f"{inst} — {INSTRUMENT_NAMES[inst]}  各合约行情")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # 表头第一行：合约名
    ws.cell(row=2, column=1, value="日期").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    col = 2
    for code in inst_codes:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 4)
        hcell = ws.cell(row=2, column=col, value=code)
        hcell.font = HEADER_FONT
        hcell.fill = PatternFill(start_color=COLORS.get(inst, "666666"), end_color=COLORS.get(inst, "666666"), fill_type="solid")
        hcell.alignment = Alignment(horizontal="center")
        for c in range(col, col + 5):
            ws.cell(row=2, column=c).border = THIN_BORDER
        col += 5

    # 表头第二行：字段名
    ws.cell(row=3, column=1, value="").border = THIN_BORDER
    col = 2
    for code in inst_codes:
        for field in ["开盘", "最高", "最低", "收盘", "结算"]:
            cell = ws.cell(row=3, column=col, value=field)
            cell.font = Font(name="微软雅黑", bold=True, size=9)
            cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            col += 1

    # 数据行
    for row_idx, d in enumerate(all_dates, 4):
        date_cell = ws.cell(row=row_idx, column=1, value=d)
        date_cell.font = DATA_FONT
        date_cell.border = THIN_BORDER
        date_cell.alignment = Alignment(horizontal="center")

        col = 2
        for code in inst_codes:
            df = all_data[code].set_index("date")
            if d in df.index:
                row = df.loc[d]
                vals = [
                    round(float(row["open"]), 1),
                    round(float(row["high"]), 1),
                    round(float(row["low"]), 1),
                    round(float(row["close"]), 1),
                    round(float(row["settle"]), 1) if row.get("settle") and float(row["settle"]) != 0 else None,
                ]
            else:
                vals = [None] * 5

            for v in vals:
                cell = ws.cell(row=row_idx, column=col)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="right")
                if v is not None:
                    cell.value = v
                col += 1

    # 列宽
    ws.column_dimensions["A"].width = 12
    for col in range(2, len(inst_codes) * 5 + 2):
        ws.column_dimensions[get_column_letter(col)].width = 9

    return ws


def build_volume_oi_sheet(wb, all_data, inst):
    """构建成交量和持仓量 sheet。"""
    ws = wb.create_sheet(title=f"{inst} 量仓")

    inst_codes = sorted([c for c in all_data if c.startswith(inst)], key=lambda c: c[2:])
    if not inst_codes:
        return ws

    all_dates = set()
    for code in inst_codes:
        all_dates.update(all_data[code]["date"].tolist())
    all_dates = sorted(all_dates)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(inst_codes) * 2 + 1)
    ws.cell(row=1, column=1, value=f"{inst} — {INSTRUMENT_NAMES[inst]}  成交量 / 持仓量").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="日期").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    col = 2
    for code in inst_codes:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        hcell = ws.cell(row=2, column=col, value=code)
        hcell.font = HEADER_FONT
        hcell.fill = PatternFill(start_color=COLORS.get(inst, "666666"), end_color=COLORS.get(inst, "666666"), fill_type="solid")
        hcell.alignment = Alignment(horizontal="center")
        for c in range(col, col + 2):
            ws.cell(row=2, column=c).border = THIN_BORDER
        col += 2

    ws.cell(row=3, column=1, value="").border = THIN_BORDER
    col = 2
    for code in inst_codes:
        for field in ["量", "仓"]:
            cell = ws.cell(row=3, column=col, value=field)
            cell.font = Font(name="微软雅黑", bold=True, size=9)
            cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            col += 1

    for row_idx, d in enumerate(all_dates, 4):
        ws.cell(row=row_idx, column=1, value=d).font = DATA_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

        col = 2
        for code in inst_codes:
            df = all_data[code].set_index("date")
            if d in df.index:
                row = df.loc[d]
                for field in ["volume", "hold"]:
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = DATA_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="right")
                    try:
                        cell.value = int(float(row[field]))
                        cell.number_format = "#,##0"
                    except (ValueError, TypeError):
                        pass
                    col += 1
            else:
                for _ in range(2):
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER
                    col += 1

    ws.column_dimensions["A"].width = 12
    for col in range(2, len(inst_codes) * 2 + 2):
        ws.column_dimensions[get_column_letter(col)].width = 11

    return ws


def build_etf_sheet(wb, start_date, end_date, create_sheet=True):
    """构建 ETF 收盘价 sheet（或仅获取数据）。"""
    ws = wb.create_sheet(title="ETF收盘价") if create_sheet else None

    etf_map = {
        "510050": "IH_上证50ETF",
        "510300": "IF_沪深300ETF",
        "510500": "IC_中证500ETF",
        "512100": "IM_中证1000ETF",
    }

    if _has_akshare():
        import akshare as ak  # noqa: F811
    else:
        if ws is not None:
            ws.cell(row=1, column=1, value="需要 akshare")
        return ws, {}

    all_dates = set()
    etf_data = {}
    for code, name in etf_map.items():
        df = None

        # 主: akshare fund_etf_hist_em（收盘后即时可用，需 YYYYMMDD 格式）
        try:
            if _has_akshare():
                import akshare as ak
                s = start_date.replace("-", "")
                e = end_date.replace("-", "")
                df = ak.fund_etf_hist_em(symbol=code, period="daily", start_date=s, end_date=e, adjust="")
                if df is not None and not df.empty:
                    df = df.rename(columns={"日期": "date", "收盘": "close"})
                    df["date"] = df["date"].astype(str)
        except Exception:
            pass

        # 备: 新浪 K 线接口（稳定但不一定有当天数据）
        if df is None or df.empty:
            try:
                from urllib.request import Request, urlopen
                import json as _json

                sina_code = f"sh{code}"
                url = (
                    "https://money.finance.sina.com.cn/quotes_service/api/json_v2.php/"
                    f"CN_MarketData.getKLineData?symbol={sina_code}&scale=240&ma=no&datalen=200"
                )
                req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urlopen(req, timeout=15) as resp:
                    raw = _json.loads(resp.read().decode())

                rows = []
                for item in raw:
                    d = item.get("day", "")
                    if start_date <= d <= end_date:
                        rows.append({
                            "date": d,
                            "close": float(item["close"]),
                            "open": float(item["open"]),
                            "high": float(item["high"]),
                            "low": float(item["low"]),
                            "volume": float(item["volume"]),
                        })
                if rows:
                    import pandas as pd
                    df = pd.DataFrame(rows)
            except Exception:
                pass

        if df is not None and not df.empty:
            all_dates.update(df["date"].tolist())
            etf_data[code] = df.set_index("date")
            print(f"  [OK] ETF {code} ({name}): {len(df)} 条 ({df['date'].iloc[0]} ~ {df['date'].iloc[-1]})")
        else:
            print(f"  [ERR] ETF {code}: 所有数据源均失败")

    # 补充当日 ETF 数据：fund_etf_hist_em 收盘后延迟较长，用 Sina 实时行情补
    from datetime import date as _dt
    today_str = _dt.today().strftime("%Y-%m-%d")
    if today_str not in all_dates:
        try:
            from urllib.request import Request, urlopen
            sina_map = {"510050": "sh510050", "510300": "sh510300", "510500": "sh510500", "512100": "sh512100"}
            url = "https://hq.sinajs.cn/list=" + ",".join(sina_map.values())
            req = Request(url, headers={"Referer": "https://finance.sina.com.cn/", "User-Agent": "Mozilla/5.0"})
            with urlopen(req, timeout=15) as resp:
                text = resp.read().decode("gbk", errors="replace")
            import re
            rev = {v: k for k, v in sina_map.items()}
            import pandas as pd
            for m in re.finditer(r'var hq_str_(\w+)="([^"]*)"', text):
                s_code = m.group(1)
                if s_code not in rev:
                    continue
                vals = m.group(2).split(",")
                if len(vals) < 4:
                    continue
                etf_code = rev[s_code]
                try:
                    close_v = float(vals[3])
                except (ValueError, TypeError):
                    continue
                if close_v <= 0:
                    continue
                if etf_code not in etf_data or etf_data[etf_code] is None or etf_data[etf_code].empty:
                    new_df = pd.DataFrame([{"date": today_str, "close": close_v}]).set_index("date")
                    etf_data[etf_code] = new_df
                else:
                    # 追加到现有 DataFrame
                    new_row = pd.DataFrame([{"date": today_str, "close": close_v, "open": close_v, "high": close_v, "low": close_v, "volume": 0}]).set_index("date")
                    etf_data[etf_code] = pd.concat([etf_data[etf_code], new_row])
                all_dates.add(today_str)
                name = {"510050": "IH_上证50ETF", "510300": "IF_沪深300ETF", "510500": "IC_中证500ETF", "512100": "IM_中证1000ETF"}
                print(f"  [+] ETF {etf_code} ({name.get(etf_code, '')}): 补充当日 {today_str} (Sina realtime) close={close_v}")
        except Exception as e:
            print(f"  [!] ETF 当日补丁失败: {e}")

    all_dates = sorted(all_dates)

    if ws is not None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(etf_data) + 1)
        ws.cell(row=1, column=1, value=f"对应 ETF 收盘价  |  {start_date} ~ {end_date}").font = TITLE_FONT
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=2, column=1, value="日期").font = HEADER_FONT
        ws.cell(row=2, column=1).fill = HEADER_FILL
        ws.cell(row=2, column=1).border = THIN_BORDER
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        etf_order = ["510050", "510300", "510500", "512100"]
        for col, code in enumerate(etf_order, 2):
            cell = ws.cell(row=2, column=col, value=f"{code} ({etf_map[code]})")
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        for row_idx, d in enumerate(all_dates, 3):
            ws.cell(row=row_idx, column=1, value=d).font = DATA_FONT
            ws.cell(row=row_idx, column=1).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

            for col, code in enumerate(etf_order, 2):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="right")
                if code in etf_data and d in etf_data[code].index:
                    try:
                        cell.value = round(float(etf_data[code].loc[d, "close"]), 3)
                    except Exception:
                        pass

        ws.column_dimensions["A"].width = 12
        for col in range(2, len(etf_order) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 18

    return ws, etf_data


def _third_friday(year, month):
    """返回指定年月的第三个周五（股指期货交割日）。"""
    first = date(year, month, 1)
    # 周五 weekday=4，计算到第一个周五的天数
    days = (4 - first.weekday()) % 7
    return first + timedelta(days=days + 14)


def _get_third_fridays(start_str, end_str):
    """返回 start~end 范围内所有第三个周五的日期字符串集合。"""
    s = date.fromisoformat(start_str)
    e = date.fromisoformat(end_str)
    result = set()
    y, m = s.year, s.month
    while True:
        tf = _third_friday(y, m)
        if tf > e:
            break
        if tf >= s:
            result.add(tf.isoformat())
        m += 1
        if m > 12:
            m = 1
            y += 1
    return result


def _compute_ratio_data(all_data, etf_sheet_data):
    """计算所有品种当月合约的期货/ETF比率。返回 (ratio_data, dates)。"""
    etf_map = {
        "IH": "510050",
        "IF": "510300",
        "IC": "510500",
        "IM": "512100",
    }

    # 收集所有期货交易日
    all_fut_dates = set()
    for df in all_data.values():
        all_fut_dates.update(df["date"].tolist())
    all_fut_dates = sorted(all_fut_dates)

    # 预处理 ETF close
    etf_close = {}
    for code, df in etf_sheet_data.items():
        if df is not None and not df.empty:
            for d, row in df.iterrows():
                k = str(d)
                if k not in etf_close:
                    etf_close[k] = {}
                try:
                    etf_close[k][code] = float(row["close"])
                except Exception:
                    pass

    # 预处理期货
    fut_by_inst_date = {}
    for code, df in all_data.items():
        inst = code[:2]
        month = code[2:]
        if inst not in fut_by_inst_date:
            fut_by_inst_date[inst] = {}
        for _, row in df.iterrows():
            d = str(row["date"])
            try:
                close = float(row["close"])
            except Exception:
                continue
            if d not in fut_by_inst_date[inst]:
                fut_by_inst_date[inst][d] = []
            fut_by_inst_date[inst][d].append((month, close))

    # 算比率
    ratio_data = {}
    for d in all_fut_dates:
        ratio_data[d] = {}
        for inst in INSTRUMENTS:
            contracts = fut_by_inst_date.get(inst, {}).get(d, [])
            if not contracts:
                continue
            contracts.sort(key=lambda x: x[0])
            front_close = contracts[0][1]
            etf_code = etf_map[inst]
            e_close = etf_close.get(d, {}).get(etf_code)
            if e_close and e_close > 0:
                ratio_data[d][inst] = round(front_close / e_close, 1)

    dates = sorted(ratio_data.keys())
    return ratio_data, dates


def _build_single_ratio_sheet(wb, ratio_data, dates, inst, start_date, end_date):
    """为单个品种构建「期货/ETF比率」sheet，含折线图 + 交割日标注。"""
    instr_names = {"IH": "上证50", "IF": "沪深300", "IC": "中证500", "IM": "中证1000"}
    chart_colors = {
        "IH": "4472C4",
        "IF": "ED7D31",
        "IC": "70AD47",
        "IM": "FFC000",
    }

    ws = wb.create_sheet(title=f"{inst}_期货ETF比率")
    name = instr_names.get(inst, inst)

    # ── 计算交割日（含假日顺延） ──
    delivery_dates = _get_third_fridays(start_date, end_date)
    # 若第三个周五为非交易日，顺延到下一个交易日
    effective_delivery = set()
    dates_set = set(dates)
    for dd in sorted(delivery_dates):
        if dd in dates_set:
            effective_delivery.add(dd)
        else:
            # 找到 dates 中紧接其后的第一个交易日
            from datetime import timedelta
            dd_obj = date.fromisoformat(dd)
            for offset in range(1, 10):
                candidate = (dd_obj + timedelta(days=offset)).isoformat()
                if candidate in dates_set:
                    effective_delivery.add(candidate)
                    break

    # ── 标题 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title_cell = ws.cell(row=1, column=1, value=f"{inst} ({name}) 当月合约 期货/ETF 比率  |  {start_date} ~ {end_date}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # ── 表头 ──
    for col, label in [(1, "日期"), (2, f"{inst} 期货/ETF")]:
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # ── 数据行 ──
    delivery_idxs = []  # 交割日在 dates 中的索引
    for row_idx, d in enumerate(dates, 3):
        ws.cell(row=row_idx, column=1, value=d).font = DATA_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

        cell = ws.cell(row=row_idx, column=2)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = "0.0"
        if inst in ratio_data[d]:
            cell.value = ratio_data[d][inst]

        if d in effective_delivery and inst in ratio_data[d]:
            delivery_idxs.append(row_idx - 3)

    # ── 列宽 ──
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16

    # ── 折线图 ──
    chart = LineChart()
    chart.title = f"{inst} ({name}) 期货/ETF 比率"
    chart.style = 10
    chart.width = 26
    chart.height = 15
    chart.y_axis.title = "比率"

    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1
    chart.x_axis.numFmt = "m/d"
    chart.x_axis.delete = False

    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(dates))
    data_ref = Reference(ws, min_col=2, min_row=2, max_row=2 + len(dates))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    s = chart.series[0]
    s.graphicalProperties.line.width = 28000
    if inst in chart_colors:
        s.graphicalProperties.line.solidFill = chart_colors[inst]

    # 所有点加微小同色圆点（确保每个点都能悬停）
    from openpyxl.chart.marker import Marker
    color = chart_colors.get(inst, "4472C4")
    s.marker.symbol = "circle"
    s.marker.size = 4
    s.marker.graphicalProperties.solidFill = color

    # 交割日点覆盖为大红色圆点
    for idx in delivery_idxs:
        pt = DataPoint(idx=idx)
        m = Marker(symbol="circle", size=9)
        m.graphicalProperties.solidFill = "FF0000"
        pt.marker = m
        s.data_points.append(pt)

    chart.legend.position = "b"
    ws.add_chart(chart, "D3")

    return ws


def build_ratio_chart_sheets(wb, all_data, etf_sheet_data, start_date, end_date):
    """为四个品种各建一个「期货/ETF比率」sheet，每个含折线图。"""
    ratio_data, dates = _compute_ratio_data(all_data, etf_sheet_data)
    if not ratio_data:
        ws = wb.create_sheet(title="期货ETF比率")
        ws.cell(row=1, column=1, value="无数据")
        return {}

    for inst in INSTRUMENTS:
        _build_single_ratio_sheet(wb, ratio_data, dates, inst, start_date, end_date)
        print(f"  [OK] {inst}_期货ETF比率: {len(dates)} 个交易日, 图表已嵌入")

    return ratio_data


def main():
    parser = argparse.ArgumentParser(description="导出 CFFEX 股指期货历史数据到 Excel")
    parser.add_argument("--start", default="2026-06-01", help="起始日期 YYYY-MM-DD")
    parser.add_argument("--end", default=date.today().isoformat(), help="结束日期 YYYY-MM-DD")
    parser.add_argument("--output", default=None, help="输出 Excel 文件路径")
    args = parser.parse_args()

    start_date = args.start
    end_date = args.end

    if not _has_akshare():
        print("❌ 需要 akshare，请先安装: pip install akshare")
        return 1

    print(f"\n📊 发现活跃合约...")
    contracts = discover_active_contracts(min_last_date=start_date)
    print(f"   共 {len(contracts)} 个合约: {', '.join(contracts)}\n")

    print(f"📥 拉取 {start_date} ~ {end_date} 历史数据...")
    all_data = fetch_all_history(contracts, start_date, end_date)

    if not all_data:
        print("❌ 没有任何数据")
        return 1

    # 输出文件名
    if args.output:
        output_path = args.output
    else:
        output_path = f"cffex_futures_{start_date}_{end_date}.xlsx"

    print(f"\n📝 生成 Excel: {output_path}")
    wb = Workbook()

    # 获取 ETF 数据（不建 sheet）
    print(f"\n📥 拉取 ETF 数据...")
    _, etf_data = build_etf_sheet(wb, start_date, end_date, create_sheet=False)

    # 删除默认空白 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 期货/ETF 比率图表（每个品种独立 sheet）
    print(f"\n📈 生成期货/ETF比率图表（每品种独立）...")
    build_ratio_chart_sheets(wb, all_data, etf_data, start_date, end_date)

    wb.save(output_path)
    print(f"\n✅ 已保存: {output_path}")
    print(f"   共 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")
    return output_path


if __name__ == "__main__":
    main()
